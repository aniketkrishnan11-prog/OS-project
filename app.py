from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Excel file path
EXCEL_FILE = 'form_data.xlsx'

# Initialize Excel file if it doesn't exist
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Form Data"
        # Add headers
        ws.append(["Name", "Email", "Phone", "Message"])
        wb.save(EXCEL_FILE)

@app.route("/", methods=["GET", "POST"])
def home():
    message = None
    if request.method == "POST":
        # Get form data
        name = request.form.get("name")
        email = request.form.get("email")
        phone = request.form.get("phone")
        user_message = request.form.get("message")
        
        # Save to Excel
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.append([name, email, phone, user_message])
            wb.save(EXCEL_FILE)
            message = "✓ Your data has been saved successfully!"
        except Exception as e:
            message = f"Error saving data: {str(e)}"
    
    return render_template("home.html", message=message)

@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/contact")
def contact():
    return render_template("contact.html")

if __name__ == "__main__":
    init_excel()  # Create Excel file on startup
    app.run(debug=True, host='0.0.0.0')
